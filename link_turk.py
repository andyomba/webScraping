import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from prefect import flow, task
import asyncio

# utils
from utils.requests import request_sync
from utils.logs import logger_step
from utils.pipelines import Pipeline
import utils.solicitudes as sl
from utils.db_actions import delete_table, insert_into_table_by_df
from utils.scrapers import unify_dfs
import utils.scrapers as scr

# flows
from subflows.set_estado import set_prefect_state_scraper

# params
queryInsert = """INSERT INTO "data_lake_approvals" ("brand_name_drug","file_number","api","holder","status","approval_date","observations","compound_name","lote","pais",
                    "molecula",
                    "link_pdf",
                    "holder_country",
                    "local_distributor",
                    "ff_manufacturer",
                    "ff_manufacturer_country",
                    "api_manufacturer",
                    "api_manufacturer_country",
                    "packaging",
                    "packaging_country",
                    "dosage_form",
                    "strength",
                    "unit",
                    "applicant_date",
                    "commercialization_date",
                    "on_the_market",
                    "expire_date",
                    "link_registro",
                    "search_value") VALUES %s"""
table_name = 'data_lake_approvals'
pais = "Turkey1_1_approvals"

@task(log_prints=True)
def scrape_data(context):
    lote = context['paisLote'].lote
    logger = context['logger']
    logger.write('INFO', 'Inicia Turkey1_1_approvals')
    
    # URLs de las páginas
    urls = "https://www.titck.gov.tr/dinamikmodul/85"

    # Headers para simular la solicitud de un navegador
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'es-ES,es;q=0.9,en;q=0.8',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36'
    }

    # Función para obtener el primer enlace de Excel y descargarlo
    def get_and_download_excel_link(url, download_folder='downloads'):
        # Inicializar el DataFrame combinado
        combined_df = pd.DataFrame()

        # Verificar si la carpeta de descargas existe, y si es así, eliminar todos los archivos dentro
        if os.path.exists(download_folder):
            for file_name in os.listdir(download_folder):
                file_path = os.path.join(download_folder, file_name)
                if os.path.isfile(file_path):
                    os.remove(file_path)
            logger.write('INFO',f"Archivos anteriores eliminados en la carpeta {download_folder}.")
        else:
            os.makedirs(download_folder)
        
        try:
            # Realizar la solicitud HTTP
            response = request_sync('GET',url, headers=headers)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Buscar todos los enlaces en la página
            links = soup.find_all('a', href=True)
            
            # Filtrar enlaces que terminen en .xls, .xlsx o .csv
            for link in links:
                href = link['href']
                if href.endswith(('.xls', '.xlsx', '.csv')):
                    # Construir la URL completa (en caso de que sea relativa)
                    if href.startswith('http'):
                        file_url = href
                    else:
                        file_url = url + href if href.startswith('/') else os.path.join(url, href)
                    
                    # Obtener el nombre del archivo
                    file_name = os.path.basename(file_url)
                    
                    # Descargar el archivo
                    file_path = os.path.join(download_folder, file_name)
                    logger.write('INFO',f"Descargando {file_url} a {file_path}...")
                    file_response = request_sync('GET',file_url)
                    with open(file_path, 'wb') as f:
                        f.write(file_response.content)
                    
                    # Leer el archivo Excel en un DataFrame de pandas
                    if file_name.endswith('.csv'):
                        # Si el archivo es CSV, lo leemos como un CSV
                        sheet_df = pd.read_csv(file_path)
                    else:
                        # Si es un archivo Excel, lo leemos usando pandas
                        sheets = pd.read_excel(file_path, sheet_name=None)
                        wb = load_workbook(file_path)

                        # Procesar las hojas específicas dependiendo de la URL
                        if url == "https://www.titck.gov.tr/dinamikmodul/85":
                            for sheet_name, sheet_df in sheets.items():
                                sheet_df = sheet_df.applymap(lambda x: x.replace('\n', ' ') if isinstance(x, str) else x)  # Limpiar los saltos de línea
                                
                                sheet_df.columns = sheet_df.iloc[0]  # Primera fila como nombres de columna
                                
                                sheet_df = sheet_df[1:]  # Eliminar la primera fila que contiene los nombres originales

                                sheet_df.columns = sheet_df.columns.str.strip().str.lower()

                                # Renombrar las columnas específicas usando un diccionario
                                sheet_df = sheet_df.rename(columns={ 
                                    'etki̇n madde': 'api', 
                                    'ürün adi': 'brand_name_drug', 
                                    'ruhsat numarasi': 'file_number', 
                                    'ruhsat sahi̇bi̇': 'holder', 
                                    'ruhsat tari̇hi̇': 'approval_date',
                                    'barkod':'observations'
                                })
                                
                                # Asignar los valores de la columna 'api' a la columna 'compound_name'
                                sheet_df['compound_name'] = sheet_df['api']
                                
                                # Si la hoja corresponde a 'RUHSATLI ÜRÜNLER LİSTESİ', revisamos el color de la fila
                                if sheet_name == 'RUHSATLI ÜRÜNLER LİSTESİ':
                                    # Usar un bucle para iterar por las filas de la hoja
                                    colors = []
                                    for row in wb[sheet_name].iter_rows(min_row=3, max_row=wb[sheet_name].max_row, min_col=2, max_col=2):
                                        # Obtener el color de la primera celda de la fila
                                        row_cells = row[0]  # Solo la primera celda de cada fila
                                        color_index = row_cells.fill.start_color.index
                                        colors.append(color_index)

                                    # Crear un array de status basado en los colores
                                    status_array = ['suspended' if color == 'FF0000' else 'approved' for color in colors]

                                    # Asignar directamente los valores de 'status' en el DataFrame
                                    sheet_df['status'] = status_array

                                    # Eliminar columnas no deseadas
                                    columns_to_keep = ["brand_name_drug", "file_number", "api", "holder", "status","approval_date","observations", "compound_name"]
                                    sheet_df = sheet_df[columns_to_keep]
                                    
                                    # Concatenar el DataFrame de la hoja procesada al DataFrame combinado
                                    combined_df = pd.concat([combined_df, sheet_df], ignore_index=True)
                                    
                                    break

                    # Después de procesar el primer archivo, salir del ciclo
                    break  # Esto asegura que solo se descargue y procese el primer archivo Excel encontrado

        except requests.exceptions.RequestException as e:
            raise ValueError(f"Error en la solicitud: {e}")
        except Exception as e:
            raise ValueError(f"Error al procesar el archivo Excel: {e}")
        
        return combined_df  # Retornar el DataFrame combinado con las hojas procesadas

    # Llamar a la función para obtener y procesar los datos
    combined_df = get_and_download_excel_link(urls)
    
    combined_df['lote'] = lote
    combined_df['pais'] = pais
    
    columns_empty = [
                    "molecula",
                    "link_pdf",
                    "holder_country",
                    "local_distributor",
                    "ff_manufacturer",
                    "ff_manufacturer_country",
                    "api_manufacturer",
                    "api_manufacturer_country",
                    "packaging",
                    "packaging_country",
                    "dosage_form",
                    "strength",
                    "unit",
                    "applicant_date",
                    "commercialization_date",
                    "on_the_market",
                    "expire_date",
                    "link_registro",
                    "search_value"
                ]
    for column in columns_empty:
        combined_df[column] = None
    
    combined_df = combined_df.applymap(lambda x: x.lower() if isinstance(x, str) else x)  # Convertir a minúsculas
    
    context['df'] = combined_df
    
    logger.write('INFO', 'Finaliza el script')
    
    return context

@flow(log_prints=True)
async def turkey1_1_approvals(paisLote):
    context = {
        'paisLote': paisLote, 
        'logger': None, 
        'query_insert': queryInsert, 
        'table_name': table_name, 
        'flow_status': None, 
        'df': None,
        'pais': pais,
        'drop_duplicates_columns': [],
    }

    pipeline = Pipeline(
        ('logger', logger_step),
        ('process_scrape_data', scrape_data),
        ('delete_table', delete_table),
        ('insert_into_table_by_df', insert_into_table_by_df),
    )

    context = await pipeline(context)


if __name__ == "__main__":
    asyncio.run(turkey1_1_approvals(sl.PaisLote(lote=764, pais='Turkey1_1_approvals', usuarioRpa='10', tipoBusqueda='api')))


